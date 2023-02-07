# coding=utf-8
# =============================================
# @Time      : 2022-03-15 15:17
# @Author    : DongWei1998
# @FileName  : 训练数据构建.py
# @Software  : PyCharm
# =============================================
from tqdm import tqdm
import openpyxl
import os


def ner_data():
    # 新数据格式
    text_train = open('../datasets/text_train.txt', 'w', encoding='utf-8')
    labels_train = open('../datasets/labels_train.txt', 'w', encoding='utf-8')
    text_val = open('../datasets/text_val.txt', 'w', encoding='utf-8')
    labels_val = open('../datasets/labels_val.txt', 'w', encoding='utf-8')
    text_test = open('../datasets/text_test.txt', 'w', encoding='utf-8')
    labels_test = open('../datasets/labels_test.txt', 'w', encoding='utf-8')

    # 旧数据
    train_data = open('../data_old/example.train', 'r', encoding='utf-8').readlines()
    dev_data = open('../data_old/example.dev', 'r', encoding='utf-8').readlines()
    test_data = open('../data_old/example.test', 'r', encoding='utf-8').readlines()
    all_data = train_data + dev_data + test_data

    with tqdm(range(len(all_data)), desc='ner 格式化') as tbar:
        for idx, str_t in enumerate(all_data):
            # print(idx
            str_t = str_t.replace('\n', '')
            if len(str_t) > 1:
                t, l = str_t.split(' ')
                if idx <= len(all_data) * 0.7:
                    text_train.write(t + ' ')
                    labels_train.write(l + ' ')
                    if t == '。':
                        text_train.write('\n')
                        labels_train.write('\n')
                elif len(all_data) * 0.7 < idx <= len(all_data) * 0.9:
                    text_val.write(t + ' ')
                    labels_val.write(l + ' ')
                    if t == '。':
                        text_val.write('\n')
                        labels_val.write('\n')
                else:
                    text_test.write(t + ' ')
                    labels_test.write(l + ' ')
                    if t == '。':
                        text_test.write('\n')
                        labels_test.write('\n')
            tbar.update(1)


def res_data():
    # 旧数据
    excel_data = openpyxl.load_workbook('../original_data/all_flzsk_data_20210608.xlsx')
    table_names = excel_data.sheetnames
    table = excel_data[table_names[2]]
    max_row = table.max_row
    with tqdm(range(max_row), desc='fw_res_data 格式化') as tbar:
        for idx in range(2, max_row):
            text = table.cell(idx, 1).value
            label = table.cell(idx, 3).value
            file_ = os.path.join(f'./fw_res_data',f'{label}')
            if os.path.exists(file_):
                with open(f"{file_}/{idx}.txt",'w',encoding='utf-8') as r:
                    r.write(text.strip())
            else:
                os.mkdir(file_)
                with open(f"{file_}/{idx}.txt",'w',encoding='utf-8') as r:
                    r.write(text.strip())
            tbar.update(1)


def sim_data():
    import jieba
    jieba.load_userdict(open('../config/count_dice_falv.txt', 'r', encoding='utf-8').readlines())
    o_data = openpyxl.load_workbook('all_fawu_data.xlsx')
    names = o_data.sheetnames
    table = o_data[names[0]]
    max_row = table.max_row
    with tqdm(desc="数据转换", total=max_row) as bar:
        for i in range(1, max_row):
            text = table.cell(i, 12).value
            try:
                text = ''.join(text.replace('\n', '').replace(' ', ''))
            except Exception as e:
                print(e)
                continue
            text_l = ' '.join([w for w in jieba.cut(text)])
            with open(f'./sim_data/{i}.txt','w',encoding='utf-8') as w:
                w.write(text_l)
            bar.update(1)

# sim_data()


def data_classfile_generate():
    # 创建文件夹
    def mkdrs(fall_path):
        if not os.path.exists(fall_path):
            os.makedirs(fall_path)
        return fall_path
    # 训练、测试、验证
    parent_directory = '../datasets'
    f_name_list = ['train', 'val', 'test']
    text_list = []
    label_list = []
    for file in ['../datasets/SENTI_ROBUST/dev.tsv',
                 '../datasets/SENTI_ROBUST/train.tsv']:
        with open(file,'r',encoding='utf-8') as r:
            infos = r.readlines()[1:]
            for info in infos:
                if file.endswith('dev.tsv'):
                    qid,label,text_a = info.replace('\n','').split('\t')
                else:
                    label, text_a = info.replace('\n','').split('\t')
                text_list.append(text_a)
                if label == '0':
                    label_list.append('消极')
                else:
                    label_list.append('积极')



    if len(text_list) == len(label_list):
        print(f'开始生成训练数据！！！ text: {len(text_list)} vs  label: {len(label_list)}')
        for split_num,text in enumerate(text_list):
            label = label_list[split_num]
            if split_num <= len(text_list) * 0.7:
                with open(os.path.join(mkdrs(os.path.join(parent_directory, f_name_list[0], f'{label}')),
                                       f'{split_num}.txt'), 'w', encoding='utf-8') as w:
                    w.write(text)
            elif len(text_list) * 0.7 < split_num <= len(text_list) * 0.9:
                with open(os.path.join(mkdrs(os.path.join(parent_directory, f_name_list[1], f'{label}')),
                                       f'{split_num}.txt'), 'w', encoding='utf-8') as w:
                    w.write(text)
            else:
                with open(os.path.join(mkdrs(os.path.join(parent_directory, f_name_list[2], f'{label}')),
                                       f'{split_num}.txt'), 'w', encoding='utf-8') as w:
                    w.write(text)
    else:
        raise print('数据于标签长度不一！！！')


data_classfile_generate()