# coding=utf-8
# =============================================
# @Time      : 2022-04-06 17:08
# @Author    : DongWei1998
# @FileName  : data_help.py
# @Software  : PyCharm
# =============================================
import os
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
import json
from utils import parameter
from utils import tokenization
import tensorflow as tf
import numpy as np
import openpyxl
from tqdm import tqdm
import jieba

# 加载停用词表
stop_word= [word.replace('\n','') for word in open('./config/stopwords.txt','r',encoding='utf-8').readlines()]
# 加载自定义词典
# jieba.load_userdict('./config/custom_dictionary.txt')



# 数据读取 【test label】
def _read_data(input_file):
    with open(input_file, 'r', encoding='utf-8') as f:
        lines = []
        words = []
        labels = []
        label_n = 4
        label_2_id = {"X":1, "[CLS]":2, "[SEP]":3}
        for line in f:
            contends = line.strip()
            tokens = contends.split(' ')
            if len(tokens) == 2:
                words.append(tokens[0])
                labels.append(tokens[1])
                if tokens[1] not in label_2_id.keys():
                    label_2_id[tokens[1]] = label_n
                    label_n += 1
            else:
                if len(contends) == 0:
                    l = ' '.join([label for label in labels if len(label) > 0])
                    w = ' '.join([word for word in words if len(word) > 0])
                    lines.append([l, w])
                    words = []
                    labels = []
                    continue
        return lines,label_2_id



class InputExample(object):
    """A single training/test example for simple sequence classification."""

    def __init__(self, guid=None, text=None, label=None):
        """Constructs a InputExample.
        Args:
          guid: Unique id for the example.
          text_a: string. The untokenized text of the first sequence. For single
            sequence tasks, only this sequence must be specified.
          label: (Optional) string. The label of the example. This should be
            specified for train and dev examples, but not for test examples.
        """
        self.guid = guid
        self.text = text
        self.label = label


class InputFeatures(object):
    """A single set of features of data."""

    def __init__(self, input_ids, input_mask, segment_ids, label_ids, ):
        self.input_ids = input_ids
        self.input_mask = input_mask
        self.segment_ids = segment_ids
        self.label_ids = label_ids
        # self.label_mask = label_mask

def _create_example(X,Y, set_type):
    examples = []
    if len(X) == len(Y):
        for i in range(len(X)):
            guid = "%s-%s" % (set_type, i)
            text = tokenization.convert_to_unicode(X[i])
            label = tokenization.convert_to_unicode(Y[i])
            examples.append(InputExample(guid=guid, text=text, label=label))
    else:
        raise print('Data length mismatch !!!')

    return examples

# 文本序列化
def filed_based_convert_examples_to_features(examples, label_2_id, max_seq_length, tokenizer):
    label_data = []
    data = []
    # 遍历训练数据
    for (ex_index, example) in enumerate(examples):
        textlist = example.text.split(' ')
        label = example.label.split(' ')
        tokens = []
        for i, word in enumerate(textlist):
            # 分词，如果是中文，就是分字,但是对于一些不在BERT的vocab.txt中得字符会被进行WordPice处理（例如中文的引号），可以将所有的分字操作替换为list(input)
            token = tokenizer.tokenize(word)
            tokens.extend(token)
        # 序列截断
        if len(tokens) >= max_seq_length - 1:
            tokens = tokens[0:(max_seq_length - 2)]  # -2 的原因是因为序列需要加一个句首和句尾标志
        ntokens = []
        segment_ids = []
        label_ids = []
        ntokens.append("[CLS]")  # 句子开始设置CLS 标志
        segment_ids.append(0)
        # append("O") or append("[CLS]") not sure!
        for i, token in enumerate(tokens):
            ntokens.append(token)
            segment_ids.append(0)
        ntokens.append("[SEP]")  # 句尾添加[SEP] 标志
        segment_ids.append(0)
        input_ids = tokenizer.convert_tokens_to_ids(ntokens)  # 将序列中的字(ntokens)转化为ID形式
        input_mask = [1] * len(input_ids)
        # label_mask = [1] * len(input_ids)
        # padding, 使用
        while len(input_ids) < max_seq_length:
            input_ids.append(0)
            input_mask.append(0)
            segment_ids.append(0)
            # we don't concerned about it!
            ntokens.append("**NULL**")
            # label_mask.append(0)
        # print(len(input_ids))
        assert len(input_ids) == max_seq_length
        assert len(input_mask) == max_seq_length
        assert len(segment_ids) == max_seq_length

        data.append(input_ids)
        label_data.append(label_2_id[label[0]])
    return data,label_data








def data_classfiles(args,set_type,mode):
    X = []
    Y = []

    if mode == 'train':
        data_file = os.path.join(args.data_dir,'train')
    elif mode == 'dev':
        data_file = os.path.join(args.data_dir,'val')
    elif mode == 'test':
        data_file = os.path.join(args.data_dir,'test')
    else:
        raise args.logger.info('mode value is not in [train dev test] ')

    # 标签映射
    if not os.path.exists(args.label_2_id_dir):
        label_2_id = {}
        # 构建标签映射
        for i, label in enumerate(os.listdir(data_file)):
            if label not in label_2_id.keys():
                label_2_id[label] = i
        with open(args.label_2_id_dir, 'w', encoding='utf-8') as w:
            w.write(json.dumps(label_2_id))
    else:
        with open(args.label_2_id_dir,'r',encoding='utf-8') as r:
            label_2_id = json.loads(r.read())

    for path_name in label_2_id.keys():
        files_path = os.path.join(data_file, path_name)
        for file in os.listdir(files_path):
            with open(os.path.join(files_path,file), 'r', encoding='utf-8') as r:
                X.append(r.read())
                Y.append(path_name)

    # 文本序列化工具
    tokenizer = tokenization.FullTokenizer(vocab_file=args.vocab_file)
    # 构建序列对象
    examples = _create_example(X,Y, set_type)
    # 文本数据序列化
    data, label = filed_based_convert_examples_to_features(examples, label_2_id, args.max_seq_length, tokenizer)
    # 将数据转换为dataset格式
    train_dataset = tf.data.Dataset.from_tensor_slices((data, label))
    # 对数据进行打乱 批次话 buffer_size 总数据量   batch 批次大小
    # fixme
    train_dataset = train_dataset.shuffle(buffer_size=len(data)).batch(args.batch_size)

    # 参数更新
    with open(args.vocab_file,'r',encoding='utf-8') as r:
        input_vocab_size = r.readlines()
    args.input_vocab_size = len(input_vocab_size)
    args.num_calss = len(label_2_id)

    return train_dataset, args


# 错别字以及大小写转换
def word_number(txt):
    number_dice = {
        '壹':1,'贰':2,'叁':3,'肆':4,'伍':5,'陆':6,'柒':7,'捌':8,'玖':9,'零':0,'一': 1,'二': 2,'三': 3,'四': 4,'五': 5,'六': 6,'七': 7,'八': 8,'九': 9,'十': 10,'百': 100,'幺': 1,'俩': 2,'两': 2,'g': 'G','块':'元','兆':'M'
    }
    text_list= []
    for word in txt:
        if word in number_dice:
            word = number_dice[word]
        text_list.append(str(word))
    return text_list

# 组合数字
def dispose_number(text_list):
    lc = ['100', '10']
    l1 = []
    l2 = []
    result = ''
    flag = False
    for i in text_list:
        if i.isdigit():         # 检测字符串是否只由数字组成
            flag = True
            if i in lc:
                if l1:
                    l2.append(int(l1[0]) * int(i))
                else:
                    l2.append(int(i))
                l1.clear()
            else:
                l1.append(i)
        else:
            if flag:    # 是否合并列表中的数字
                if l1:
                    if len(l1) > 8:
                        result = result + ''.join(l1) + i
                    else:
                        result = result + str(sum(l2) + int(l1[0])) + i
                else:
                    result = result + str(sum(list(set(l2)))) + i   # 去除重复的数字 例如 200，200g  ==》 200g
                    # result = result + str(sum(l2)) + i
                l1.clear()
                l2.clear()
                flag = False
            else:
                result = result + i
    return result

# 停用词以及标点符号过滤
def stop_word_conve(txt):
    word_list = [word for word in jieba.lcut(txt) if word not in stop_word]
    return ''.join(word_list)



def data_cleaning(data_path):
    data = openpyxl.load_workbook(data_path)    # 加载excel文件
    table_name = data.sheetnames                # 获取所有的sheet
    table = data[table_name[0]]                 # 选择第n个sheet
    rows = table.max_row                        # 获得行数
    ncolumns = table.max_column                 # 获得列数
    for row in tqdm(range(2,rows)):
        text = table.cell(row,1).value
        if text is None:
            break
        # print(text)
        label = table.cell(row,2).value
        # 数字转换
        text = word_number(text)
        # 数字合并
        text = dispose_number(text)
        # 停用词过滤
        text = stop_word_conve(text)
        # 将清洗后的数据写入excel中
        if len(text) > 10:
            table.cell(row, 6).value = text
    data.save('../data/国庆投诉数据_1.xlsx')


def create_data(file_path):
    o_data = openpyxl.load_workbook(file_path)
    names = o_data.sheetnames
    print(names)
    for name in names:
        if name=='非降档数据4200':
            label = '非降档'
        elif name=='降档数据4969':
            label = '降档'
        else:
            continue
        dataset_path = os.path.join('../datasets/val',label)
        if not os.path.exists(dataset_path):
            os.mkdir(dataset_path)
        table = o_data[name]
        max_row = table.max_row
        max_column = table.max_column
        for row in range(1,max_row+1):
            if row > 500:
                # 获取数据
                text = table.cell(row,1).value
                # 数字转换
                text = word_number(text)
                # # 数字合并
                text = dispose_number(text)
                # 停用词过滤
                text = stop_word_conve(text)
                with open(os.path.join(dataset_path, f'{row}.txt'),'w',encoding='utf-8') as w:
                    w.write(text)
            if row == 1000:
                break




if __name__ == '__main__':
    # alphamind_read_data(text_data_file='../datasets_old/text_val.txt', label_data_file='../datasets_old/labels_val.txt')

    # args = parameter.parser_opt('train')

    # train_dataset = data_set(args, file='../datasets_old/example.dev', set_type=True)

    # data_classfiles(args, set_type=True, mode='dev')

    # # 数据可视化
    # total = 0
    # for times,item in enumerate(train_dataset,1):
    #     # 遍历训练数据，相当于一个epoch
    #     if times < 20 :
    #         print(f'=======当前批数：{times} ========')
    #         print(item)
    #         print(item[0].shape)
    #     batch_count = item[0].shape[0]   # batch_size设置为512,但是不一定能整除,实际最后一个batch达不到512
    #     total += batch_count
    #     times += 1
    #
    # print('扫过数据数量:', total)


    file_path = '../config/all_data_20210907.xlsx'
    create_data(file_path)
